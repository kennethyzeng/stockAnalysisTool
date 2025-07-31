from openpyxl import load_workbook
import pandas as pd 
import numpy as np 
#https://www.geeksforgeeks.org/exporting-a-pandas-dataframe-to-an-excel-file/
#https://openpyxl.readthedocs.io/en/latest/api/openpyxl.workbook.workbook.html

########1 print sheetname#########
import openpyxl
def print_sheetname(): 
    # input excel file path
    inputExcelFile ="sampleTutorialsPoint.xlsx"

    # creating or loading an excel workbook
    newWorkbook = openpyxl.load_workbook(inputExcelFile)

    # printing all the sheetnames in an excel file using sheetnames attribute
    print('The Sheet names of the given excel file: ')

    # Getting the sheetnames as a list using the sheetnames attribute
    sheetNames=newWorkbook.sheetnames

    # Traversing in the sheetNames list
    for name in sheetNames:
        print(name)

######## 2 append dataframe data with creation of new sheetname at exisitng excel#####
#####Normal handling of to_excel will override excel data#####
#link: https://github.com/pandas-dev/pandas/issues/33264
#link:https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas
#https://saturncloud.io/blog/how-to-add-a-worksheet-to-an-existing-excel-file-with-pandas/
#class pandas.ExcelWriter(path, engine=None, date_format=None, datetime_format=None, mode='w', storage_options=None, if_sheet_exists=None, engine_kwargs=None, **kwargs)

path='smci.xlsx'
writer = pd.ExcelWriter(path, engine = 'openpyxl', mode='a',if_sheet_exists='replace')


x3 = np.random.randn(100, 2)
df3 = pd.DataFrame(x3)

x4 = np.random.randn(100, 2)
df4 = pd.DataFrame(x4)

df3.to_excel(writer, sheet_name = 'x3', index=False)
df4.to_excel(writer, sheet_name = 'x4', index=False)
writer.close()       

#######3 insert data or modify data to specific sheet at excel 
#https://www.excel-learn.com/insert-columns-openplxl/
#https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
#https://www.geeksforgeeks.org/working-with-excel-spreadsheets-in-python/
#to change specific sheet, need to swtich active sheet number, then modify data. default active # is 0
wb=load_workbook('smci.xlsx')
wb.active=2
sheet = wb.active 
#wb.autofit(axis="columns")  auto change column length, but need to verify
#https://gist.github.com/summerofgeorge/96dac94293b60c70d11d7cd7e852ffd6
sheet.insert_cols(idx=2)
wb.save('smci.xlsx')

#the link to set width with workbook 
#https://docs.aspose.com/cells/java/adjusting-row-height-and-column-width-in-python/#:~:text=Row%20Height%20Successfully.%22-,Setting%20the%20Column%20Width,width%2C%20the%20desired%20column%20width.


#####4 auto adjust the length of column for the sheet 
####This code didn't work 
path='smci.xlsx'
writer = pd.ExcelWriter(path, engine = 'openpyxl', mode='a',if_sheet_exists='replace') #xlsxwriter
df3  = pd.DataFrame({
    'Name': ['John', 'Jane', 'Bob', 'Alice a Long Name'], # this is a very long name
    'Age': [25, 30, 35, 40],
    'Salary': [50000, 60000, 70000, 80000]
})
df3.to_excel(writer, sheet_name = 'x3', index=True)
writer.close()

def auto_width_columns(df, sheetname):
        workbook = writer.book  
        worksheet= writer.sheets[sheetname] 
    
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).str.len().max(), len(col) + 2)
            worksheet.set_column(i, i, column_len)

#########
####
import xlsxwriter
from datetime import datetime

workbook = xlsxwriter.Workbook('datetime_bug.xlsx')
worksheet = workbook.add_worksheet()

# using the xlsxwriter functions works as intended
worksheet.write('A1', 0.05)
worksheet.write('B1', datetime(year=2020, month=1, day=1))

percentage_format = workbook.add_format({'num_format':'0.00%'})
date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})

worksheet.set_column('A:A', width=None, cell_format=percentage_format)
worksheet.set_column('B:B', width=None, cell_format=date_format)

workbook.close()