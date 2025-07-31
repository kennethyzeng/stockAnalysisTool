

from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Rename the default sheet to "Summary"
summary_sheet = wb.active
summary_sheet.title = "Summary"

# Create other sheets
price_data_sheet = wb.create_sheet("PriceData")
analysis_sheet = wb.create_sheet("Analysis")
finance_analysis_sheet = wb.create_sheet("Finance_analysis")

# Input "Current_Volumn" in cell A2 of "Analysis"
analysis_sheet["A2"] = "Current_Volumn"

# Save the workbook
wb.save("stock_report.xlsx")