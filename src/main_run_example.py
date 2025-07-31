#! usr/env/bin/python3 

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import parameter_data as data_file
from single_stock import SingleStock as ss 


stocks_list = data_file.read_stocklist_to_set()
def generate_excel_files_for_stocks(stocks_list):
    """
    read stock symbol from set format; Create excel file and worksheet for each stock systmobl, 
    then generate information for related cells
    """
    output_dir = "StockAnalysisFiles"
    os.makedirs(output_dir, exist_ok=True)

    for symbol in stocks_list:
        stock_dir = os.path.join(output_dir, symbol)
        os.makedirs(stock_dir, exist_ok=True)
        wb = Workbook()
        
        # Rename default sheet to Summary
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
      
        price_data_sheet = wb.create_sheet("PriceData")
        analysis_sheet = wb.create_sheet("Analysis")
        finance_analysis_sheet = wb.create_sheet("Finance_Analysis")
        

        #finance analysis parameter setup
        finance_analysis_parameter = data_file.single_stock_excel_finance_analysis_parameters_dicts
        for key, item in finance_analysis_parameter.items():
            if key != "column_dimensions":
                finance_analysis_sheet[key] = item[0]
                color_fill = PatternFill(start_color=item[1],  #hex code for color
                         end_color=item[1],
                         fill_type="solid")
                finance_analysis_sheet[key].fill = color_fill
            else:
                for i in item:    
                    finance_analysis_sheet.column_dimensions[i[0]].width = i[1]

       
        filename = os.path.join(stock_dir, f"{symbol}.xlsx")
        wb.save(filename)

    print(f"Generated {len(stocks_list)} Excel files in '{output_dir}'")


def data_filled_at_excel_for_each_sock(self):
    res_ss=ss("nvda")
    print(res_ss.ticker)

    h=res_ss.download_data()
    print(h)
    #pass

#        res_ss= SingleStock("nvda")
 #   print(res_ss.ticker)

  #  res_dd=res_ss.download_data()
#################
#generate_excel_files_for_stocks(stocks_list)
data_filled_at_excel_for_each_sock("smci")

